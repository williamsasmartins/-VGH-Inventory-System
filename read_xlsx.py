import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\Williams Martins\VGH Material\VGH_Pricing_Dryco_Updated.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n=== Sheet: {name} (rows={ws.max_row}, cols={ws.max_column}) ===')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 30: break
        if any(v is not None for v in row):
            # Truncate each cell to 40 chars for readability
            cells = [str(v)[:40] if v is not None else None for v in row]
            print(cells)
